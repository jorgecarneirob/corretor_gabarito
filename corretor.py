import cv2
import numpy as np
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd
from multiprocessing import Pool, cpu_count
import logging

# Configurações de logging para o módulo corretor
logging.basicConfig(level=logging.INFO)

# ======= PARÂMETROS DO FORMULÁRIO PADRÃO =======
CANVAS_W, CANVAS_H = 674, 790
GRID_ROWS, GRID_COLS = 18, 9
ID_ROW = 4
PROVA_ROW = 5
ID_COLS = range(2, 8)
PROVA_COLS = range(2, 6)
ANS_ROWS = range(7, 17)
ANS_COLS = range(3, 8)

# ---------- PRÉ-PROCESSAMENTO ----------
def load_and_preprocess_image(image_path):
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        logging.error(f"Erro: Não consegui abrir a imagem: {image_path}")
        return None
    
    blur = cv2.GaussianBlur(img, (5, 5), 0)
    _, binary = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    k = np.ones((3, 3), np.uint8)
    binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, k, iterations=1)
    binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, k, iterations=1)
    return binary

# ---------- DETECÇÃO ESCALA-INVARIANTE DOS 4 MARCADORES ----------
def find_marker_centers(binary):
    target_ar = 70.0 / 46.0
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    H, W = binary.shape[:2]
    img_area = W * H

    cand = []
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        if w < 5 or h < 5:
            continue
        area = cv2.contourArea(cnt)
        bbox_area = w * h
        if bbox_area == 0:
            continue
        rectangularity = area / float(bbox_area)
        ratio = max(w, h) / float(min(w, h))

        if area < img_area * 0.0003 or area > img_area * 0.05:
            continue
        if rectangularity < 0.75:
            continue
        if abs(ratio - target_ar) > 0.35:
            continue

        cx, cy = x + w / 2.0, y + h / 2.0
        cand.append(((cx, cy), area))

    if len(cand) < 4:
        logging.error(f"Erro: apenas {len(cand)} candidatos a marcador encontrados. Esperado: 4.")
        return None

    cand = sorted(cand, key=lambda t: t[1], reverse=True)[:10]
    pts = np.array([c[0] for c in cand], dtype=np.float32)

    s = pts[:, 0] + pts[:, 1]
    d = pts[:, 0] - pts[:, 1]
    tl = pts[np.argmin(s)]
    br = pts[np.argmax(s)]
    tr = pts[np.argmax(d)]
    bl = pts[np.argmin(d)]

    centers = [tuple(tl), tuple(tr), tuple(bl), tuple(br)]
    return centers

# ---------- WARP POR PERSPECTIVA ----------
def warp_to_standard(binary, centers, width=CANVAS_W, height=CANVAS_H, save_debug=False, debug_prefix="debug"):
    pts_src = np.float32(centers)
    pts_dst = np.float32([
        [0, 0],
        [width - 1, 0],
        [0, height - 1],
        [width - 1, height - 1]
    ])
    M = cv2.getPerspectiveTransform(pts_src, pts_dst)
    warped = cv2.warpPerspective(binary, M, (width, height))

    if save_debug:
        cv2.imwrite(f"{debug_prefix}_warp.png", warped)
    return warped

# ---------- GRADE A PARTIR DE 4 PONTOS (CENTROS) ----------
def compute_grid_centers_from_points(corners_pts, rows=GRID_ROWS, cols=GRID_COLS):
    tl = np.array(corners_pts[0], dtype=np.float32)
    tr = np.array(corners_pts[1], dtype=np.float32)
    bl = np.array(corners_pts[2], dtype=np.float32)
    br = np.array(corners_pts[3], dtype=np.float32)

    grid_centers = []
    for i in range(rows):
        left = tl + (bl - tl) * (i / (rows - 1))
        right = tr + (br - tr) * (i / (rows - 1))
        row = []
        for j in range(cols):
            p = left + (right - left) * (j / (cols - 1))
            row.append((int(round(p[0])), int(round(p[1]))))
        grid_centers.append(row)
    return grid_centers

# ---------- LEITURAS ----------
def read_binary_value(binary, points):
    bits = []
    H, W = binary.shape[:2]
    for x, y in points:
        x = int(x); y = int(y)
        x0 = max(x - 2, 0); x1 = min(x + 3, W)
        y0 = max(y - 2, 0); y1 = min(y + 3, H)
        area = binary[y0:y1, x0:x1]
        mean = float(np.mean(area))
        bits.append('1' if mean > 127 else '0')
    return int(''.join(bits), 2)

def read_answers(binary, grid_centers, debug=False, debug_prefix="debug"):
    answers = []
    debug_image = cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)

    for q_idx, row in enumerate(ANS_ROWS):
        max_mean = -1.0
        selected_letter = "-"
        for i, col in enumerate(ANS_COLS):
            x, y = grid_centers[row][col]
            x0 = max(x - 2, 0); x1 = min(x + 3, binary.shape[1])
            y0 = max(y - 2, 0); y1 = min(y + 3, binary.shape[0])
            area = binary[y0:y1, x0:x1]
            mean = float(np.mean(area))
            if mean > max_mean:
                max_mean = mean
                selected_letter = chr(65 + i)
            if debug:
                cv2.circle(debug_image, (x, y), 5, (255, 0, 0), -1)
                cv2.putText(debug_image, chr(65 + i), (x - 10, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 0, 0), 1)

        answers.append(selected_letter if max_mean >= 10 else "-")

    if debug:
        cv2.imwrite(f"{debug_prefix}_respostas.png", debug_image)
    return answers

# ---------- PIPELINE DE UMA IMAGEM ----------
def process_gabarito(image_path, save_debug=False):
    binary0 = load_and_preprocess_image(image_path)
    if binary0 is None:
        return None
    
    centers = find_marker_centers(binary0)
    if centers is None:
        return None

    warped = warp_to_standard(binary0, centers, width=CANVAS_W, height=CANVAS_H,
                              save_debug=save_debug, debug_prefix=os.path.basename(image_path).split('.')[0])
    
    corners_dst = [(0, 0), (CANVAS_W - 1, 0), (0, CANVAS_H - 1), (CANVAS_W - 1, CANVAS_H - 1)]
    grid_centers = compute_grid_centers_from_points(corners_dst, GRID_ROWS, GRID_COLS)

    id_points = [grid_centers[ID_ROW][c] for c in ID_COLS]
    prova_points = [grid_centers[PROVA_ROW][c] for c in PROVA_COLS]

    student_id = read_binary_value(warped, id_points)
    prova_id = read_binary_value(warped, prova_points)
    respostas = read_answers(warped, grid_centers, debug=save_debug, debug_prefix=os.path.basename(image_path).split('.')[0])
    
    return {
        "Aluno_ID": student_id,
        "Prova_ID": prova_id,
        "Respostas": respostas,
        "Caminho_Arquivo": image_path
    }

# ---------- LEITURA DO GABARITO TXT ----------
def load_gabarito(filepath):
    gabarito = {}
    if not os.path.exists(filepath):
        logging.error(f"Erro: Arquivo de gabarito não encontrado em {filepath}")
        return None
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            prova, questao, peso_q, pesos_alt_str, correta = line.split("|")
            peso_q = float(peso_q)
            pesos_alt = {k: float(v) for k, v in (pair.split(":") for pair in pesos_alt_str.split(","))}
            gabarito.setdefault(prova, {})[questao] = {
                "peso_questao": peso_q,
                "pesos_alternativas": pesos_alt,
                "correta": correta
            }
    return gabarito

# ---------- EXPORTAÇÃO EXCEL ----------
def export_to_excel(resultados, gabarito, filename="resultados.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    if not resultados:
        logging.warning("Não há resultados para exportar.")
        return

    sample_gabarito_data = next(iter(gabarito.values()))
    questoes = sorted(list(sample_gabarito_data.keys()), key=lambda x: int(x.strip("Q ")))

    headers = ["Aluno", "Prova"] + questoes + ["Total"]
    ws.append(headers)

    vermelho_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    verde_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    amarelo_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    for prova_id in sorted(gabarito.keys(), key=lambda x: (len(x), x)):
        prova_gabarito = gabarito[prova_id]
        gabarito_row = [f"GABARITO_P{prova_id}", ""]
        for q in questoes:
            gabarito_row.append(prova_gabarito.get(q, {}).get("correta", "-"))
        gabarito_row.append("")
        ws.append(gabarito_row)

    for r in resultados:
        aluno_id = r.get("Aluno_ID")
        prova_id = str(r.get("Prova_ID"))
        respostas = r.get("Respostas", [])

        prova_gabarito = gabarito.get(prova_id, {})
        total = 0.0
        row = [aluno_id, prova_id]

        for i, questao in enumerate(questoes):
            resp = respostas[i] if i < len(respostas) else "-"
            info = prova_gabarito.get(questao)
            if info:
                peso_q = info["peso_questao"]
                peso_alt = info["pesos_alternativas"].get(resp, 0.0)
                total += peso_q * peso_alt
                row.append(resp)
            else:
                row.append("-")

        row.append(total)
        ws.append(row)

        for i, questao in enumerate(questoes):
            cell = ws.cell(row=ws.max_row, column=i + 3)
            info = prova_gabarito.get(questao, {})
            correta = info.get("correta")
            pesos_alt = info.get("pesos_alternativas", {})
            resp = respostas[i] if i < len(respostas) else "-"

            if correta:
                if resp == correta:
                    peso_alt = pesos_alt.get(resp, 0.0)
                    if abs(peso_alt - 1.0) < 1e-6:
                        cell.fill = verde_fill
                    elif 0.1 < peso_alt < 1.0:
                        cell.fill = amarelo_fill
                    else:
                        cell.fill = vermelho_fill
                else:
                    peso_alt = pesos_alt.get(resp, 0.0)
                    cell.fill = amarelo_fill if peso_alt > 0.0 else vermelho_fill
            else:
                cell.fill = vermelho_fill

    wb.save(filename)
    logging.info(f"Arquivo {filename} salvo com sucesso.")
    return filename

# --- FUNÇÃO PRINCIPAL PARA SER CHAMADA PELO APP.PY ---
def run_correction_parallel(image_paths, gabarito_data, output_excel_path):
    num_processes = min(cpu_count(), 4)
    logging.info(f"Iniciando correção de {len(image_paths)} imagens com {num_processes} processos.")
    
    tasks = [(path, gabarito_data) for path in image_paths]
    
    resultados = []
    with Pool(processes=num_processes) as pool:
        resultados = pool.starmap(process_gabarito, tasks)
        
    resultados = [r for r in resultados if r is not None]

    if not resultados:
        logging.warning("Nenhuma prova foi processada com sucesso.")
        return None

    return export_to_excel(resultados, gabarito_data, filename=output_excel_path)

# --- FUNÇÃO AUXILIAR PARA O APP.PY CHAMAR DIRETO ---
def process_single_file_and_get_results(image_path, gabarito_data):
    return process_gabarito(image_path, save_debug=False)